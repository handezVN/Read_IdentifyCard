# import xlsxwriter module
import xlsxwriter
import openpyxl
import index
# Workbook() takes one, non-optional, argument
# which is the filename that we want to create.
def export_Excel(Persons,today):
    flag = 2 ;
    while(flag == 2):
        try:
            path  = ("C:\\Users\\Handez\\Documents\\pythonProject1\\"+str(today)+".xlsx")
            # workbook object is created
            wb_obj = openpyxl.load_workbook(path)

            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            row = m_row +1 ;
            for x in Persons:
                sheet_obj.cell(row=row, column=1).value = x.hoten
                sheet_obj.cell(row=row, column=2).value = x.socd
                sheet_obj.cell(row=row, column=3).value = x.address1 + x.address2
                sheet_obj.cell(row=row, column=4).value = x.ngaysinh
                row = row + 1
            wb_obj.save(str(today)+".xlsx")
            flag = 1;
            return ;
        except:
            workbook = xlsxwriter.Workbook(str(today)+".xlsx")
            worksheet = workbook.add_worksheet()
            worksheet.write('A1', 'Họ và Tên')
            worksheet.write('B1', 'Số CCCD')
            worksheet.write('C1', 'Địa Chỉ')
            worksheet.write('D1', 'Ngày sinh')
            worksheet.write('E1', 'Ngày Đến')
            worksheet.write('F1', 'NGày đi')
            workbook.close()
# The workbook object is then used to add new
# worksheet via the add_worksheet() method.

#
# # Use the worksheet object to write
# # Header data name

#
# # Finally, close the Excel file
# # via the close() method.
